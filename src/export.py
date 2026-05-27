from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TIER_FILLS = {
    "Red":   PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "Amber": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Green": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
}

_EXPORT_COLS = [
    "sku", "store", "category", "urgency_tier", "markdown_pct",
    "stock_on_hand", "velocity_14d", "days_of_cover",
    "sell_through_pct", "action_by_date", "rationale",
]
_COL_RENAME = {
    "sku": "SKU", "store": "Store", "category": "Category",
    "urgency_tier": "Urgency Tier", "markdown_pct": "Recommended Markdown %",
    "stock_on_hand": "Current Stock", "velocity_14d": "Daily Velocity (14d)",
    "days_of_cover": "Days of Cover", "sell_through_pct": "Sell-Through Rate %",
    "action_by_date": "Action By Date", "rationale": "Rationale",
}


def build_action_sheet(recommendations_df: pd.DataFrame) -> BytesIO:
    cols = [c for c in _EXPORT_COLS if c in recommendations_df.columns]
    df = recommendations_df[cols].rename(columns=_COL_RENAME).round(2)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Recommendations")
        ws = writer.sheets["Recommendations"]

        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        tier_col_idx = list(df.columns).index("Urgency Tier") + 1
        for row_idx in range(2, ws.max_row + 1):
            tier_val = ws.cell(row=row_idx, column=tier_col_idx).value
            fill = _TIER_FILLS.get(tier_val, PatternFill())
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    buffer.seek(0)
    return buffer
