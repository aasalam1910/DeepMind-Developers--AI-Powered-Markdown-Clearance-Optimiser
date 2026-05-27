import pytest
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from src.export import build_action_sheet


def make_rec_df():
    return pd.DataFrame([
        {
            "sku": "SKU-001", "store": "MUM-1", "category": "Apparel",
            "urgency_tier": "Red", "markdown_pct": 35,
            "stock_on_hand": 100, "velocity_14d": 1.5,
            "days_of_cover": 66.7, "sell_through_pct": 45.0,
            "action_by_date": "2026-05-30", "rationale": "Test rationale.",
        },
        {
            "sku": "SKU-002", "store": "DEL-2", "category": "Footwear",
            "urgency_tier": "Green", "markdown_pct": 0,
            "stock_on_hand": 30, "velocity_14d": 4.0,
            "days_of_cover": 7.5, "sell_through_pct": 85.0,
            "action_by_date": None, "rationale": "",
        },
    ])


def test_returns_bytes_io():
    assert isinstance(build_action_sheet(make_rec_df()), BytesIO)


def test_has_correct_column_headers():
    wb = load_workbook(build_action_sheet(make_rec_df()))
    headers = [cell.value for cell in wb.active[1]]
    assert "SKU" in headers
    assert "Urgency Tier" in headers
    assert "Recommended Markdown %" in headers
    assert "Rationale" in headers


def test_correct_row_count():
    wb = load_workbook(build_action_sheet(make_rec_df()))
    assert wb.active.max_row == 3  # header + 2 data rows


def test_header_has_dark_fill():
    wb = load_workbook(build_action_sheet(make_rec_df()))
    header_cell = wb.active["A1"]
    assert header_cell.fill.fgColor.rgb.upper() in ("FF1E293B", "1E293B")
